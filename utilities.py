import csv
import os
import yaml
import cv2
import numpy as np
from icecream import ic
import openpyxl
from ultralytics import YOLO
import re
from transliterate import translit
import polars as pl
import openpyxl
import csv

import io
import fitz
from PIL import Image
from paddleocr import PaddleOCR
from pdf2image import convert_from_path
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import utilities


def read_classes(classes_file_path: str = "classes.txt") -> list[str]:
    classes = []
    with open(classes_file_path, "r") as recognized_text_file:
        for line in recognized_text_file.readlines():
            if line.strip():
                classes.append(line.strip())
    return classes


def read_image(file_path: str) -> np.ndarray | None:
    image_file_name = os.path.basename(file_path)
    image_file_name, image_file_extension = os.path.splitext(image_file_name)

    if image_file_extension in [".png", ".jpg"]:
        image = cv2.imread(file_path, cv2.IMREAD_COLOR)

        return image

    elif image_file_extension == ".pdf":
        pdf_read = fitz.open(file_path)
        for page in pdf_read:
            pix = page.get_pixmap(
                matrix=fitz.Identity,
                dpi=None,
                colorspace=fitz.csRGB,
                clip=None,
                alpha=True,
                annots=True,
            )
            image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                pix.height, pix.width, pix.n
            )

            return image


def load_model(model_file_path: str) -> YOLO:
    model = YOLO(model_file_path)
    return model


def get_predictions(model: YOLO, image: np.ndarray) -> list:
    predictions = model(image)[0]

    boxes = [[int(coord) for coord in box.xyxy[0].tolist()] for box in  predictions.boxes]
    class_indices = [int(box.cls[0]) for box in  predictions.boxes]
    confidence_scores = [int(box.conf[0]) for box in predictions.boxes]

    return {
        "boxes": boxes,
        "class_indices": class_indices,
        "confidence_scores": confidence_scores,
    }


def bind_text_boxes(
    ocr: PaddleOCR,
    image: np.ndarray,
    detection_results: list,
    shkaf_detection_results: list,
) -> list[tuple[str, tuple, tuple]]:
    """
    output format:
    [
        // shkaf name, class name, object coords, text coords
        ("ВРУ-1", "QF", (x1, y1, x2, y2), (x1, y1, x2, y2)),
        ...
    ]
    """

    # Get shkaf names and coords
    shkaf_coords: dict[str, tuple] = {}
    for shkaf_index, shkaf_box in enumerate(shkaf_detection_results["boxes"]):
        shkaf_text = f"SHKAF{shkaf_index + 1}"
        
        shkaf_coords[shkaf_text] = (
            int(shkaf_box[0]),
            int(shkaf_box[1]),
            int(shkaf_box[2]),
            int(shkaf_box[3]),
        )

    ic(shkaf_coords)

    elements = []
    texts = {}

    classes_list = []
    with open("config.yaml") as config_file:
        config = yaml.safe_load(config_file)
        classes_path = config["data"]["classes_path"]
        classes_list=utilities.read_classes(classes_path)
        
    
    # Разделение результатов на элементы и тексты
    for box, box_class_idx in zip(detection_results["boxes"], detection_results["class_indices"]):
        class_name = classes_list[box_class_idx]
        x, y, w, h = box[0], box[1], box[2] - box[0], box[3] - box[1]

        if class_name.endswith("_TEXT"):
            if class_name not in texts:
                texts[class_name] = []
            texts[class_name].append({"bbox": (x, y, w, h)})
        else:
            elements.append({"class": class_name, "bbox": (x, y, w, h)})
            
    def calculate_distance(bbox1, bbox2):
        x1, y1, w1, h1 = bbox1
        x2, y2, w2, h2 = bbox2
        center1 = (x1 + w1 / 2, y1 + h1 / 2)
        center2 = (x2 + w2 / 2, y2 + h2 / 2)
        return np.sqrt((center1[0] - center2[0]) ** 2 + (center1[1] - center2[1]) ** 2)

    near_elements = []
    for element in elements:
        el_class = element["class"]
        el_bbox = element["bbox"]
        text_class = f"{el_class}_TEXT"

        element_shkaf_name = ""

        # select corresponding shkaf
        for shkaf_name, coords in shkaf_coords.items():
            el_x, el_y, el_w, el_h = el_bbox

            shkaf_x1, shkaf_y1, shkaf_x2, shkaf_y2 = coords

            # Проверяем, находится ли центр el_bbox внутри coords
            if (shkaf_x1 <= el_x + el_w / 2 <= shkaf_x2) and (
                shkaf_y1 <= el_y + el_h / 2 <= shkaf_y2
            ):
                element_shkaf_name = shkaf_name
                break

        if text_class in texts:
            nearest_text = min(
                texts[text_class], key=lambda t: calculate_distance(el_bbox, t["bbox"])
            )
            el_x, el_y, el_w, el_h = el_bbox

            txt_x, txt_y, txt_w, txt_h = nearest_text["bbox"]
            near_elements.append(
                (
                    element_shkaf_name,
                    el_class,
                    (el_x, el_y, el_x + el_w, el_y + el_h),
                    (txt_x, txt_y, txt_x + txt_w, txt_y + txt_h),
                )
            )

    return near_elements


# def find_shkaf_text(detection_shkaf: list) -> list[tuple[str, tuple, tuple]]:

#     results = detection_results
#     elements = []
#     text = {}

#     for shkaf_box in results.b

#     cls_shkaf = int(box_shkaf.cls)


def prepare_dataframe(dataframe_path: str):
    try:
        df = pl.read_excel(dataframe_path)

        df_ = df.select(
            [
                "__UNNAMED__0",
                "Дата актуальности:",
                "__UNNAMED__11",
            ]
        ).rename(
            {
                "__UNNAMED__0": "Артикул",
                "Дата актуальности:": "Номенклатура",
                "__UNNAMED__11": "Стоимость",
            }
        )

        df_n = df_.slice(11, df_.height - 11)

        return df_n

    except Exception as e:
        print(f"Error while processing dataframe: {str(e)}")
        return


def recognize_paddle_ocr(ocr: PaddleOCR, region: np.ndarray) -> str:
    def extract_valid_texts(results) -> list[str]:
        """
        Извлечение и фильтрация текста из результатов OCR.
        """
        valid_texts = []
        if results is not None:
            for res in results:
                if isinstance(res, list) and res:
                    for line in res:
                        if (
                            isinstance(line, list)
                            and len(line) > 1
                            and isinstance(line[1], tuple)
                        ):
                            text = line[1][0].strip()
                            if any(char.isalpha() for char in text) and any(
                                char.isdigit() for char in text
                            ):
                                valid_texts.append(text)
        return valid_texts

    def rotate_image(img: np.ndarray, angle: int) -> np.ndarray:
        """
        Поворачивает изображение на заданный угол.
        """
        return np.rot90(img, k=angle // 90)

    def calculate_confidence(result):
        """
        Вычисляет общую уверенность для результата OCR.
        """
        confidence = 0
        if result is not None:
            for res in result:
                if isinstance(res, list) and res:
                    for line in res:
                        if (
                            isinstance(line, list)
                            and len(line) > 1
                            and isinstance(line[1], tuple)
                            and len(line[1]) > 1
                        ):
                            confidence += line[1][1]
        return confidence

    best_result = []
    best_confidence = 0

    for angle in [0, 90, 180, 270]:
        rotated_region = rotate_image(region, angle)
        result = ocr.ocr(rotated_region, cls=True)
        valid_texts = extract_valid_texts(result)

        confidence = calculate_confidence(result)

        if confidence > best_confidence:
            best_result = valid_texts
            best_confidence = confidence

    return " ".join(best_result)


def recognize_text(
    ocr: PaddleOCR,
    image: np.ndarray,
    binded_results: list[tuple[str, tuple]],
) -> list[tuple[str, str]]:
    """
    input format:
        [
            // shkaf name,  class name, text coords
            ("ВРУ-1", "QF", (x1, y1, x2, y2)),
            ...
        ]

    output format:
        [
            // shkaf name, class name, recognized text
            ("ВРУ-1", "QF", "iC60N (C), 1P 16A"),
            ...
        ]
    """

    recognized_results = []
    for shkaf_name, object_name, object_text_coords in binded_results:
        x1, y1, x2, y2 = map(int, object_text_coords)
        # print(type(image))
        # print(image.shape)
        # print(x1, y1, x2, y2)

        # image_region = image[y1:y2, x1:x2, :]
        image_region = image[y1:y2, x1:x2]
        recognized_text = recognize_paddle_ocr(ocr, image_region)
        recognized_results.append((shkaf_name, object_name, recognized_text))

    return recognized_results


# Нормализация текста - перевод в нижний регистр + в латиницу
def find_elements(df, search_string, abbreviations):
    def normalize(text):
        text_lower = text.lower()
        text_cyrillic = re.sub(r"[^\w\s]", "", translit(text_lower, "ru"))
        text_latin = re.sub(r"[^\w\s]", "", translit(text_lower, "ru", reversed=True))
        return text_cyrillic, text_latin

    def expand_abbreviations(text, abbr_dict):
        words = text.split()
        expanded = [abbr_dict.get(word.upper(), word) for word in words]
        return " ".join(expanded)

    def prepare_search_elements(search_string, abbreviations):
        expanded_search = expand_abbreviations(search_string, abbreviations)
        return [normalize(elem) for elem in expanded_search.split()]

    def check_match(x, search_elements):
        norms = normalize(x)
        return all(
            any(se[0] in norm or se[1] in norm for norm in norms)
            for se in search_elements
        )

    search_elements = prepare_search_elements(search_string, abbreviations)
    result = df.filter(
        pl.col("Номенклатура").map_elements(lambda x: check_match(x, search_elements))
    )
    if len(result) > 0:
        return result.sort("Стоимость").row(0)
    return None


def read_dict_from_txt(file_path: str) -> dict:
    data = {}
    with open(file_path, "r", encoding="utf-8") as file:
        for line in file:
            key, value = line.strip().split(":", 1)
            data[key] = value
    return data


def read_list_from_txt(file_path: str) -> list:
    txt_file = open(file_path, "r", encoding="utf-8")
    content = list(txt_file.read().split("\n"))
    return content


def fuzzy_match(query, choices, method="ratio", threshold=70):
    if method == "ratio":
        match_func = fuzz.ratio
    elif method == "partial_ratio":
        match_func = fuzz.partial_ratio
    elif method == "token_sort_ratio":
        match_func = fuzz.token_sort_ratio
    elif method == "token_set_ratio":
        match_func = fuzz.token_set_ratio
    else:
        raise ValueError(
            "Invalid method. Choose from 'ratio', 'partial_ratio', 'token_sort_ratio', or 'token_set_ratio'"
        )

    matches = process.extractBests(
        query, choices, scorer=match_func, score_cutoff=threshold
    )
    return matches  # This will return a list of tuples (match, score)


# Теперь мы можем использовать эту функцию в process_search_list



def process_search_list(
    search_list, df, abbreviations, threshold=70
) -> list[tuple[str, tuple]]:
    results = []
    for shkaf_name, object_name, text in search_list:
        search_string = f"{object_name} {text}"
        expanded_search = expand_abbreviations(search_string, abbreviations)
        choices = df["Номенклатура"].to_list()
        matches = process.extractBests(
            expanded_search,
            choices,
            scorer=fuzz.token_set_ratio,
            score_cutoff=threshold,
        )
        if matches:
            best_match, score = matches[0]
            corresponding_row = df.filter(pl.col("Номенклатура") == best_match).row(0)
            results.append(
                (shkaf_name, search_string, (best_match, score, corresponding_row))
            )
        else:
            results.append((shkaf_name, search_string, None))
    return results


def expand_abbreviations(text, abbr_dict):
    words = text.split()
    expanded = [abbr_dict.get(word.upper(), word) for word in words]
    return " ".join(expanded)

def create_table(filename, results, output_format="both"):
    """
    ic| find_el: [
        ('QF BA47- 1P 16A',
            ('Выключатель автоматический AV-6 1P 16A (B) 6kA EKF AVERES',
                79,
                (
                    'mcb6-1-16B-av',
                    'Выключатель автоматический AV-6 1P 16A (B) 6kA EKF AVERES',
                    '401.7216'
                )
            )
        ),
        ...
    ]
    """

    article_occurences: dict[str, int] = {}
    article_params: dict[str, tuple] = {}  # article, nomenclature, price

    for result in sorted(results, key=lambda x: x[0]):  # sort by shraf name
        if result[2]:
            shkaf_name, search_string, (best_match, score, corresponding_row) = result
            article, nomenclature, cost = corresponding_row
  
            if article not in article_occurences:
                article_occurences[article] = 1
            else:
                article_occurences[article] += 1

            if article not in article_params:
                article_params[article] = (shkaf_name, nomenclature, cost)

    data_rows = []
    headers = ["Артикул", "Номенклатура", "Стоимость", "Количество", "Общая стоимость"]

    last_shkaf: str = None

    for article in article_occurences:
        shkaf_name = article_params[article][0]
        if last_shkaf is None or last_shkaf != shkaf_name:
            data_rows.append([None, shkaf_name, None, None, None])

        quantity = article_occurences[article]

        nomenclature = article_params[article][1]
        cost = float(article_params[article][2])
        total_cost = cost * quantity

        data_rows.append([article, nomenclature, cost, quantity, total_cost])

        last_shkaf = shkaf_name

    # Подсчитываем общую стоимость всех элементов
    total_sum = sum(row[4] for row in data_rows if row[4] is not None)
    data_rows.append([None, None, None, "Общая стоимость", total_sum])

    if output_format in ["both", "excel"]:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        for row_index, row_data in enumerate(data_rows, start=2):
            for col_index, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index, value=cell_value)
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjust_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjust_width
        excel_filename = f"{filename}.xlsx"
        workbook.save(excel_filename)
        print(f"Excel файл сохранен как {excel_filename}")

    if output_format in ["both", "csv"]:
        csv_filename = f"{filename}.csv"
        with open(csv_filename, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            writer.writerows(data_rows)
        print(f"CSV файл сохранен как {csv_filename}")

    ic(data_rows)

    return article_occurences

def export_table(data_rows: list[tuple], filename:str, output_format="both"):

    def is_float(string):
        try:
            float(string)
            return True
        except ValueError:
            return False

    headers = ["Артикул", "Номенклатура", "Стоимость", "Количество", "Общая стоимость"]
    
    # Подсчитываем общую стоимость всех элементов
    total_sum = sum(float(row[4]) if is_float(row[4]) else 0 for row in data_rows if row[4] is not None)
    data_rows.append([None, None, None, "Общая стоимость", total_sum])

    if output_format in ["both", "excel"]:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        for row_index, row_data in enumerate(data_rows, start=2):
            for col_index, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_index, column=col_index, value=cell_value)
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjust_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjust_width
        excel_filename = f"{filename}.xlsx"
        workbook.save(excel_filename)
        print(f"Excel файл сохранен как {excel_filename}")

    if output_format in ["both", "csv"]:
        csv_filename = f"{filename}.csv"
        with open(csv_filename, mode="w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerow(headers)
            writer.writerows(data_rows)
        print(f"CSV файл сохранен как {csv_filename}")

    ic(data_rows)